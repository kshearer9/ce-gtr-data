"""Does soft counting estimate the field distribution better than hard labels?

RQ3 is a distributional question: what share of funded circular economy research
sits in each field, against what share of the published output does. It is not a
question about any individual project. That distinction matters, because a
classifier can be poor at the per-project task and still be usable for the
aggregate one, provided its errors are handled rather than ignored.

Three estimators are compared on the 287 gold projects, using out-of-fold
predictions from set-up H so nothing is scored on its own training data:

    hard    assign each project to its argmax field and count. This is what
            projects_labelled_final.csv currently does. It inherits the
            classifier's bias directly: a model that over-reaches for
            Engineering produces a distribution that over-states Engineering.

    soft    add each project's predicted probability to every field. A project
            the model reads as 55% Engineering and 40% Environmental Science
            contributes to both. Errors that are genuine uncertainty partly
            cancel instead of compounding.

    adjusted  soft counts corrected by the out-of-fold confusion matrix, solving
            for the class prior that would have produced the observed counts.
            This is the strongest correction available but the least stable,
            because it inverts a matrix estimated on 287 cases.

Scored by total variation distance from the true distribution, by mean absolute
error in percentage points, and per field, since the fields that matter here are
Engineering and Environmental Science specifically.

Run from the repo root, after run_variant.py --crosswalk james:

    /opt/anaconda3/bin/python scripts/classification/test_soft_counts.py
"""
from pathlib import Path
import argparse
import json
import re
import time

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.sparse import vstack as spvstack
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression

SEED = 20260803
C_TFIDF, C_EMB, MAX_ITER = 1.0, 3.0, 3000
WEIGHT = 100
PATTERN = re.compile(r"^\s*(.+?)\s*\((\d+(?:\.\d+)?)%\)\s*$")

EXTENSION = {
    "Medical & health interface": "Medicine",
    "Pol. sci. & internat. studies": "Social Sciences",
    "Atmospheric phys. & chemistry": "Earth and Planetary Sciences",
    "Social Policy": "Social Sciences",
    "Particle physics - experiment": "Physics and Astronomy",
    "Astronomy - observation": "Physics and Astronomy",
    "Optics, photonics & lasers": "Physics and Astronomy",
    "Supercond, magn. &quant.fluids": "Physics and Astronomy",
    "Linguistics": "Arts and Humanities", "Archaeology": "Arts and Humanities",
    "Particle astrophysics": "Physics and Astronomy",
    "Particle physics - theory": "Physics and Astronomy",
    "Demography & human geography": "Social Sciences",
    "Philosophy": "Arts and Humanities",
    "Planetary science": "Earth and Planetary Sciences",
    "Music": "Arts and Humanities", "Drama & theatre studies": "Arts and Humanities",
    "Area Studies": "Social Sciences", "Astronomy - theory": "Physics and Astronomy",
    "Solar & terrestrial physics": "Physics and Astronomy",
    "Nuclear physics": "Physics and Astronomy", "Plasma physics": "Physics and Astronomy",
    "Social Work": "Social Sciences", "Library & information studies": "Social Sciences",
    "Demography": "Social Sciences", "Classics": "Arts and Humanities",
    "Facility Development": None, "Dance": "Arts and Humanities",
}


def parse_subjects(value):
    out = []
    for part in str(value).split(";"):
        part = part.strip()
        if not part:
            continue
        m = PATTERN.match(part)
        out.append((m.group(1).strip(), float(m.group(2))) if m else (part, 100.0))
    return out


def load_mapping(xw):
    ws = load_workbook(xw)["Crosswalk (FINAL)"]
    mapping = {}
    for row in range(7, ws.max_row + 1):
        subject = ws.cell(row, 1).value
        if subject:
            field = ws.cell(row, 3).value
            mapping[subject] = None if field == "EXCLUDED" else field
    mapping.update(EXTENSION)
    return mapping


def primary_field(value, mapping):
    mapped = [(mapping[s], w) for s, w in parse_subjects(value) if mapping.get(s)]
    return max(mapped, key=lambda t: t[1])[0] if mapped else None


def tvd(p, q):
    """Total variation distance: half the L1 gap between two distributions."""
    return 0.5 * float(np.abs(np.asarray(p) - np.asarray(q)).sum())


def main() -> None:
    ROOT = Path(__file__).resolve().parents[2]
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=str(ROOT / "data" / "classification"))
    ap.add_argument("--crosswalk",
                    default=str(ROOT / "data" / "crosswalk" /
                                "crosswalk_gtr_to_openalex_FINAL.xlsx"))
    ap.add_argument("--projects",
                    default=str(ROOT / "data" / "cleaned" / "merged" / "projects.csv"))
    args = ap.parse_args()
    R = Path(args.root)
    OUT = ROOT / "data" / "classification" / "results"
    OUT.mkdir(parents=True, exist_ok=True)
    t0 = time.time()

    mapping = load_mapping(args.crosswalk)
    # Gold is rebuilt from projects.csv rather than read from gold_james.csv.
    # The rule is deterministic, and the staged copy of that file turned out to
    # hold an older 276-row set while the folds describe 287. Rebuilding and
    # asserting against the folds means a mismatch stops the run instead of
    # silently scoring the wrong projects.
    proj_all = pd.read_csv(args.projects, low_memory=False)
    gold = proj_all[proj_all.research_subjects.notna()].copy()
    gold["primary_field"] = gold.research_subjects.map(lambda s: primary_field(s, mapping))
    gold = gold[gold.primary_field.notna()].reset_index(drop=True)
    vc = gold.primary_field.value_counts()
    gold = gold[~gold.primary_field.isin(vc[vc < 5].index)].reset_index(drop=True)

    folds = json.load(open(R / "folds_james.json"))["folds"][:5]
    fold_ids = set(folds[0]["train"]) | set(folds[0]["test"])
    assert fold_ids == set(gold.project_id), (
        f"gold rebuild ({len(gold)}) does not match the frozen folds "
        f"({len(fold_ids)}); the crosswalk or projects.csv has changed")
    classes = sorted(gold.primary_field.unique())
    y = gold.primary_field.values
    loc = {p: i for i, p in enumerate(gold.project_id)}
    gold_text = (gold.title_clean.fillna("") + ". "
                 + gold.abstract_text_clean.fillna("")).str.strip().values

    pemb = np.load(R / "project_embeddings_mpnet.npy")
    pidx = pd.read_csv(R / "project_embedding_index.csv")
    # A stale index against a fresh embedding array is the worst failure mode
    # here, because it does not raise, it silently pairs each project with
    # another project's vector. Refuse to run rather than produce plausible
    # nonsense.
    if len(pidx) != len(pemb):
        raise SystemExit(
            f"project_embedding_index.csv has {len(pidx)} rows but "
            f"project_embeddings_mpnet.npy has {len(pemb)}. The index is stale. "
            f"Re-run embed_texts_mpnet.py --projects-only before continuing.")
    prow = {p: i for i, p in enumerate(pidx.project_id)}
    Xg = np.stack([pemb[prow[p]] for p in gold.project_id])

    corpus = pd.read_csv(R / "gtr_tagged_corpus.csv", low_memory=False)
    corpus["primary_field"] = corpus.research_subjects.map(lambda s: primary_field(s, mapping))
    cemb = np.load(R / "corpus_embeddings_mpnet.npy")
    cidx = pd.read_csv(R / "corpus_embedding_index.csv")
    corpus = corpus.set_index("project_id").reindex(cidx.project_id).reset_index()
    ce_ids = set(pd.read_csv(args.projects, usecols=["project_id"]).project_id)
    overlap = corpus.project_id.isin(ce_ids).values
    ckeep = corpus.primary_field.isin(classes).values & ~overlap
    Xc, yc = cemb[ckeep], corpus.primary_field.values[ckeep]
    ctext = (corpus.title.fillna("") + ". "
             + corpus.abstract_text.fillna("")).str.strip().values[ckeep]
    print(f"gold {len(gold)}, {len(classes)} classes | corpus in scope {ckeep.sum()} "
          f"(leakage guard dropped {int(overlap.sum())}) [{time.time()-t0:.0f}s]")

    vec = TfidfVectorizer(sublinear_tf=True, min_df=3, ngram_range=(1, 2),
                          stop_words="english", max_features=200_000)
    Zc = vec.fit_transform(ctext)
    Zg = vec.transform(gold_text)
    print(f"TF-IDF {Zc.shape[1]} features [{time.time()-t0:.0f}s]")

    # ---- out-of-fold probability matrix from set-up H ----------------------
    P = np.zeros((len(gold), len(classes)))
    for fold in folds:
        tr = [loc[p] for p in fold["train"]]
        te = [loc[p] for p in fold["test"]]
        w = np.concatenate([np.ones(len(yc)), np.full(len(tr), float(WEIGHT))])
        ytr = np.concatenate([yc, y[tr]])
        emb = LogisticRegression(max_iter=MAX_ITER, C=C_EMB, class_weight="balanced")
        emb.fit(np.vstack([Xc, Xg[tr]]), ytr, sample_weight=w)
        tf = LogisticRegression(max_iter=MAX_ITER, C=C_TFIDF, class_weight="balanced")
        tf.fit(spvstack([Zc, Zg[tr]]), ytr, sample_weight=w)
        assert list(tf.classes_) == list(emb.classes_) == classes
        P[te] = (emb.predict_proba(Xg[te]) + tf.predict_proba(Zg[te])) / 2
        print(f"  split {fold['split']} done [{time.time()-t0:.0f}s]", flush=True)
    np.save(R / "oof_proba_H.npy", P)

    idx = {c: i for i, c in enumerate(classes)}
    true_counts = np.array([(y == c).sum() for c in classes], dtype=float)
    true_dist = true_counts / true_counts.sum()

    hard_pred = np.array(classes)[np.argmax(P, axis=1)]
    hard_dist = np.array([(hard_pred == c).sum() for c in classes], dtype=float)
    hard_dist /= hard_dist.sum()

    soft_dist = P.sum(axis=0) / P.sum()

    # confusion-corrected: solve M q = observed, M[i,j] = P(predict i | true j)
    M = np.zeros((len(classes), len(classes)))
    for j, c in enumerate(classes):
        rows = P[y == c]
        if len(rows):
            M[:, j] = rows.mean(axis=0)
    try:
        adj = np.linalg.lstsq(M, soft_dist, rcond=None)[0]
        adj = np.clip(adj, 0, None)
        adj = adj / adj.sum() if adj.sum() > 0 else soft_dist
    except np.linalg.LinAlgError:
        adj = soft_dist

    out = pd.DataFrame({"field": classes, "true": true_dist, "hard": hard_dist,
                        "soft": soft_dist, "adjusted": adj})
    out["hard_err_pp"] = (out.hard - out["true"]) * 100
    out["soft_err_pp"] = (out.soft - out["true"]) * 100
    out["adj_err_pp"] = (out.adjusted - out["true"]) * 100
    out.to_csv(OUT / "soft_count_comparison.csv", index=False)

    print("\n=== share of the gold set in each field (%) ===")
    show = out.copy()
    for c in ("true", "hard", "soft", "adjusted"):
        show[c] = (show[c] * 100).round(1)
    print(show[["field", "true", "hard", "soft", "adjusted",
                "hard_err_pp", "soft_err_pp", "adj_err_pp"]]
          .round(1).sort_values("true", ascending=False).to_string(index=False))

    print("\n=== how far each estimator is from the truth ===")
    for name, dist in (("hard", hard_dist), ("soft", soft_dist), ("adjusted", adj)):
        mae = float(np.abs(dist - true_dist).mean() * 100)
        mx = float(np.abs(dist - true_dist).max() * 100)
        print(f"  {name:9s} TVD {tvd(dist, true_dist):.4f}   "
              f"mean abs error {mae:.2f}pp   worst field {mx:.2f}pp")

    print("\n=== the two fields that decide RQ3 ===")
    for c in ("Engineering", "Environmental Science"):
        if c in idx:
            i = idx[c]
            print(f"  {c:24s} true {true_dist[i]*100:5.1f}%  hard {hard_dist[i]*100:5.1f}%  "
                  f"soft {soft_dist[i]*100:5.1f}%  adjusted {adj[i]*100:5.1f}%")
    print(f"\nWrote soft_count_comparison.csv and oof_proba_H.npy. "
          f"Total {time.time()-t0:.0f}s.")


if __name__ == "__main__":
    main()
