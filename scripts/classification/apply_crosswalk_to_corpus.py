"""Map the wider GtR corpus onto OpenAlex fields using the verified crosswalk.

Applies crosswalk_gtr_to_openalex_FINAL.xlsx (the 54 subjects James verified)
plus a 28-row extension covering subjects that occur only outside the CE
subset. The extension is training-corpus only: none of these subjects appears
in the 1,380 CE projects, so it does not touch any reported CE label.

Primary field = field of the highest-weighted mapped subject, ties resolved by
first-listed order, exactly as for the CE gold standard.

Run from the repo root:
    /opt/anaconda3/bin/python scripts/classification/apply_crosswalk_to_corpus.py

Input:  data/classification/gtr_tagged_corpus.csv
        data/crosswalk/crosswalk_gtr_to_openalex_FINAL.xlsx
Output: data/classification/gtr_corpus_labelled.csv
"""
from pathlib import Path
import re
import sys

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DIR = ROOT / "data" / "classification"
SRC = DIR / "gtr_tagged_corpus.csv"
XW = ROOT / "data" / "crosswalk" / "crosswalk_gtr_to_openalex_FINAL.xlsx"

# Subjects occurring only in the wider corpus, mapped for training use.
# "Facility Development" is excluded on the same ground as "Tools, technologies
# & methods" in the verified crosswalk: it describes activity, not discipline.
EXTENSION = {
    "Medical & health interface": "Medicine",
    "Pol. sci. & internat. studies": "Social Sciences",
    "Atmospheric phys. & chemistry": "Earth and Planetary Sciences",
    "Social Policy": "Social Sciences",
    "Particle physics - experiment": "Physics and Astronomy",
    "Astronomy - observation": "Physics and Astronomy",
    "Optics, photonics & lasers": "Physics and Astronomy",
    "Supercond, magn. &quant.fluids": "Physics and Astronomy",
    "Linguistics": "Arts and Humanities",
    "Archaeology": "Arts and Humanities",
    "Particle astrophysics": "Physics and Astronomy",
    "Particle physics - theory": "Physics and Astronomy",
    "Demography & human geography": "Social Sciences",
    "Philosophy": "Arts and Humanities",
    "Planetary science": "Earth and Planetary Sciences",
    "Music": "Arts and Humanities",
    "Drama & theatre studies": "Arts and Humanities",
    "Area Studies": "Social Sciences",
    "Astronomy - theory": "Physics and Astronomy",
    "Solar & terrestrial physics": "Physics and Astronomy",
    "Nuclear physics": "Physics and Astronomy",
    "Plasma physics": "Physics and Astronomy",
    "Social Work": "Social Sciences",
    "Library & information studies": "Social Sciences",
    "Demography": "Social Sciences",
    "Classics": "Arts and Humanities",
    "Facility Development": None,
    "Dance": "Arts and Humanities",
}

PATTERN = re.compile(r"^\s*(.+?)\s*\((\d+(?:\.\d+)?)%\)\s*$")


def parse_subjects(value):
    out = []
    for part in str(value).split(";"):
        part = part.strip()
        if not part:
            continue
        m = PATTERN.match(part)
        out.append((m.group(1).strip(), float(m.group(2))) if m else (part, 100.0))
    return out


def main() -> None:
    for path in (SRC, XW):
        if not path.exists():
            sys.exit(f"Missing input: {path}")

    ws = load_workbook(XW)["Crosswalk (FINAL)"]
    verified = {}
    for row in range(7, ws.max_row + 1):
        subject = ws.cell(row, 1).value
        if subject:
            field = ws.cell(row, 3).value
            verified[subject] = None if field == "EXCLUDED" else field
    mapping = {**verified, **EXTENSION}
    print(f"Crosswalk: {len(verified)} verified subjects + {len(EXTENSION)} extension rows")

    df = pd.read_csv(SRC)
    unknown = {s for v in df.research_subjects for s, _ in parse_subjects(v)} - set(mapping)
    if unknown:
        print(f"WARNING: {len(unknown)} subjects still unmapped, e.g. {sorted(unknown)[:5]}")

    def primary(value):
        mapped = [(mapping[s], w) for s, w in parse_subjects(value) if mapping.get(s)]
        return max(mapped, key=lambda t: t[1])[0] if mapped else None

    df["primary_field"] = df.research_subjects.map(primary)
    kept = df[df.primary_field.notna()]
    out = DIR / "gtr_corpus_labelled.csv"
    df.to_csv(out, index=False)
    print(f"Mapped {len(kept)} of {len(df)} ({len(kept) / len(df):.1%}); wrote {out}")
    print("\nField distribution:")
    print(kept.primary_field.value_counts().to_string())


if __name__ == "__main__":
    main()
