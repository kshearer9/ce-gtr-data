"""Set the confidence thresholds and apply the selected classifier.

This is the step that turns the evaluation in run_variant.py into the labelled
dataset RQ3 needs. It does four things, in order:

    1. Rebuilds out-of-fold predictions from set-up H, the configuration
       selected on macro-F1, using the frozen gold set and folds that
       run_variant.py already wrote. Nothing about the model is re-chosen here.

    2. Builds an accuracy-reject curve from those out-of-fold predictions and
       reads the confidence threshold off it. The target accuracy is declared
       on the command line and recorded in the run summary, so the threshold
       is a consequence of a stated target rather than a number picked after
       seeing which one flattered the result (Hendrickx et al., 2024).

    3. Refits H on the corpus plus the whole gold set and labels every project
       that has no funder-assigned subject.

    4. Draws a stratified verification sample and writes it as a blind coding
       sheet, with the model's answers held in a separate key file that the
       coding sheet does not reference.

Three tiers come out of it:

    tier 1  funder-assigned, taken from the GtR research subject via the
            crosswalk. No model involved, so no confidence value.
    tier 2  model-assigned at or above the threshold.
    tier 3  model-assigned below the threshold. Retained and reported, but
            carrying an explicit accuracy caveat.

Run from the repo root, after run_variant.py --crosswalk james has finished:

    /opt/anaconda3/bin/python scripts/classification/apply_classifier.py

    # faster, if the variant log's "G weights chosen" line has a clear mode
    /opt/anaconda3/bin/python scripts/classification/apply_classifier.py --weight 25

The default --repeats 1 uses the first five folds, which gives each gold
project exactly one out-of-fold prediction. That keeps the accuracy-reject
curve free of the within-project correlation that pooling all 25 splits would
introduce, and it is about five times faster. --repeats 5 pools everything if
you want the smoother curve for a figure.
"""
from pathlib import Path
import argparse
import json
import re
import sys
import time

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from scipy.sparse import vstack as spvstack
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, f1_score
from sklearn.model_selection import StratifiedKFold

ROOT = Path(__file__).resolve().parents[2]
DIR = ROOT / "data" / "classification"
RESULTS = DIR / "results"
XW = ROOT / "data" / "crosswalk" / "crosswalk_gtr_to_openalex_FINAL.xlsx"
PROJECTS = ROOT / "data" / "cleaned" / "merged" / "projects.csv"
CORPUS = DIR / "gtr_tagged_corpus.csv"

SEED = 20260803
C_TFIDF, C_EMB, MAX_ITER = 1.0, 3.0, 3000
WEIGHT_GRID = [1, 5, 10, 25, 50, 100]

# Reused verbatim from run_variant.py so the crosswalk applied here cannot
# drift from the one the evaluation used.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from run_variant import load_mapping, primary_field  # noqa: E402


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def wilson(k, n, z=1.96):
    """Wilson score interval, which behaves at proportions near 1."""
    if n == 0:
        return (float("nan"), float("nan"))
    p = k / n
    d = 1 + z * z / n
    centre = (p + z * z / (2 * n)) / d
    half = z * np.sqrt(p * (1 - p) / n + z * z / (4 * n * n)) / d
    return (max(0.0, centre - half), min(1.0, centre + half))


def fit_emb(Xtr, ytr, w=None):
    clf = LogisticRegression(max_iter=MAX_ITER, C=C_EMB, class_weight="balanced")
    clf.fit(Xtr, ytr, sample_weight=w)
    return clf


def choose_weight(Xc, yc, Xg, yg):
    """Nested selection of the CE upweighting factor (Jiang and Zhai, 2007).

    run_variant.py chooses this inside every evaluation fold, which is what
    keeps its reported figures honest. Here the model is being fitted rather
    than evaluated, so the factor is chosen once on the whole gold set by
    three-fold inner cross-validation. Pass --weight to skip this.
    """
    inner = StratifiedKFold(n_splits=3, shuffle=True, random_state=7)
    best, best_score = WEIGHT_GRID[0], -1.0
    for mult in WEIGHT_GRID:
        scores = []
        for itr, ite in inner.split(Xg, yg):
            w = np.concatenate([np.ones(len(yc)), np.full(len(itr), float(mult))])
            clf = fit_emb(np.vstack([Xc, Xg[itr]]), np.concatenate([yc, yg[itr]]), w)
            scores.append(f1_score(yg[ite], clf.predict(Xg[ite]), average="macro"))
        mean = float(np.mean(scores))
        print(f"    weight x{mult:<4} macro-F1 {mean:.3f}", flush=True)
        if mean > best_score:
            best, best_score = mult, mean
    print(f"  selected weight x{best} (inner macro-F1 {best_score:.3f})")
    return best


def fit_h(Xc, yc, Zc, Xg, yg, Zg, weight):
    """Fit both halves of the set-up H ensemble on corpus plus gold."""
    w = np.concatenate([np.ones(len(yc)), np.full(len(yg), float(weight))])
    Xtr, ytr = np.vstack([Xc, Xg]), np.concatenate([yc, yg])
    emb = fit_emb(Xtr, ytr, w)
    tfidf = LogisticRegression(max_iter=MAX_ITER, C=C_TFIDF, class_weight="balanced")
    tfidf.fit(spvstack([Zc, Zg]), ytr, sample_weight=w)
    assert list(tfidf.classes_) == list(emb.classes_)
    return emb, tfidf


def predict_h(emb, tfidf, Xq, Zq):
    proba = (emb.predict_proba(Xq) + tfidf.predict_proba(Zq)) / 2
    idx = np.argmax(proba, axis=1)
    return emb.classes_[idx], proba[np.arange(len(idx)), idx], proba


# ---------------------------------------------------------------------------
# ACCURACY-REJECT CURVE
# ---------------------------------------------------------------------------

def accuracy_reject_curve(conf, correct):
    """Accuracy among retained predictions at every candidate threshold."""
    order = np.argsort(-conf)
    c_sorted, ok_sorted = conf[order], correct[order].astype(int)
    kept = np.arange(1, len(conf) + 1)
    acc = np.cumsum(ok_sorted) / kept
    rows = pd.DataFrame(dict(threshold=c_sorted, n_retained=kept,
                             coverage=kept / len(conf), accuracy=acc,
                             n_correct=np.cumsum(ok_sorted)))
    # One row per distinct threshold, keeping the largest retained set at each.
    return rows.groupby("threshold", as_index=False).last().sort_values(
        "threshold", ascending=False).reset_index(drop=True)


def pick_threshold(curve, target, min_coverage):
    """Lowest threshold whose retained accuracy still meets the target.

    Lowest, not highest, because the aim is to keep as many projects in tier 2
    as the accuracy target allows. min_coverage guards the degenerate case
    where the target is only met on a handful of very confident predictions.
    """
    ok = curve[(curve.accuracy >= target) & (curve.coverage >= min_coverage)]
    if ok.empty:
        return None
    return ok.iloc[-1]


# ---------------------------------------------------------------------------
# VERIFICATION SAMPLE
# ---------------------------------------------------------------------------

def stratified_sample(frame, n, strata, rng):
    """Proportional allocation across strata, with at least one per stratum."""
    groups = {key: block for key, block in frame.groupby(strata, observed=True)}
    sizes = pd.Series({key: len(block) for key, block in groups.items()})
    alloc = np.floor(sizes / sizes.sum() * n).astype(int).clip(lower=1)
    # Hand out or claw back the rounding remainder, largest strata first.
    order = list(sizes.sort_values(ascending=False).index)
    i = 0
    while alloc.sum() != n and i < 10_000:
        key = order[i % len(order)]
        if alloc.sum() < n and alloc[key] < sizes[key]:
            alloc[key] += 1
        elif alloc.sum() > n and alloc[key] > 1:
            alloc[key] -= 1
        i += 1
    parts = [groups[key].sample(n=min(int(take), len(groups[key])), random_state=rng)
             for key, take in alloc.items()]
    return pd.concat(parts).sample(frac=1.0, random_state=rng).reset_index(drop=True)


def coalesce(frame, name):
    """One column by that name, even when the merge left several.

    projects.csv carries duplicate column names, so frame[name] can return a
    DataFrame rather than a Series. Left alone, str() on the row value writes
    the repr of a Series into the cell, which is what happened the first time
    this sheet was built.
    """
    block = frame.loc[:, frame.columns == name]
    if block.shape[1] == 0:
        return None
    if block.shape[1] == 1:
        return block.iloc[:, 0]
    print(f"  '{name}' appears {block.shape[1]} times in projects.csv, "
          f"taking the first non-null per row")
    return block.bfill(axis=1).iloc[:, 0]


def cell(value, limit=None):
    """openpyxl refuses pandas NA, so every cell goes in as plain text."""
    if isinstance(value, pd.Series):
        nonnull = value.dropna()
        value = nonnull.iloc[0] if len(nonnull) else None
    if value is None or value is pd.NA or (isinstance(value, float) and np.isnan(value)):
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value)
    return text[:limit] if limit else text


def write_coding_sheet(sample, classes, out_xlsx):
    """A sheet with no model answer anywhere on it, so the coding is blind."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Coding"
    headers = ["sample_id", "project_id", "funder", "title", "abstract",
               "your_field", "your_confidence", "notes"]
    ws.append(headers)
    for _, r in sample.iterrows():
        ws.append([cell(r.get("sample_id")), cell(r.get("project_id")),
                   cell(r.get("funder")), cell(r.get("title_clean"), 300),
                   cell(r.get("abstract_text_clean"), 2000), "", "", ""])

    ref = wb.create_sheet("Fields")
    ref.append(["field"])
    for c in classes:
        ref.append([c])
    ref.append(["UNCLEAR"])

    conf = wb.create_sheet("Confidence")
    conf.append(["level"])
    for level in ("certain", "fairly sure", "guess"):
        conf.append([level])

    last = len(sample) + 1
    dv_field = DataValidation(type="list",
                              formula1=f"=Fields!$A$2:$A${len(classes) + 2}",
                              allow_blank=True, showDropDown=False)
    dv_conf = DataValidation(type="list", formula1="=Confidence!$A$2:$A$4",
                             allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_field)
    ws.add_data_validation(dv_conf)
    dv_field.add(f"F2:F{last}")
    dv_conf.add(f"G2:G{last}")

    widths = [11, 14, 16, 60, 90, 26, 16, 34]
    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=5):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(out_xlsx)



def _oof_and_threshold(folds, loc, gold, y, Xc, yc, Zc, Xg, Zg, weight, args, t0):
    """Out-of-fold predictions from H, then read the threshold off the curve."""
    print(f"\nout-of-fold predictions from set-up H (weight x{weight})")
    oof = []
    for fold in folds:
        tr = [loc[p] for p in fold["train"]]
        te = [loc[p] for p in fold["test"]]
        emb, tfidf = fit_h(Xc, yc, Zc, Xg[tr], y[tr], Zg[tr], weight)
        pred, conf, _ = predict_h(emb, tfidf, Xg[te], Zg[te])
        oof.append(pd.DataFrame(dict(split=fold["split"],
                                     project_id=gold.project_id.values[te],
                                     true_field=y[te], pred_field=pred,
                                     confidence=conf)))
        print(f"  split {fold['split']:2d} done [{time.time() - t0:.0f}s]", flush=True)
    oof = pd.concat(oof, ignore_index=True)
    oof["correct"] = oof.pred_field == oof.true_field
    oof.to_csv(RESULTS / "oof_predictions_H.csv", index=False)
    print(f"  overall out-of-fold accuracy {oof.correct.mean():.3f}, "
          f"macro-F1 {f1_score(oof.true_field, oof.pred_field, average='macro'):.3f}")

    curve = accuracy_reject_curve(oof.confidence.values, oof.correct.values)
    curve.to_csv(RESULTS / "accuracy_reject_curve.csv", index=False)
    print("\naccuracy-reject curve, at selected coverage points")
    for cov in (0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0):
        r = curve.iloc[(curve.coverage - cov).abs().argmin()]
        print(f"  coverage {r.coverage:.2f}  threshold {r.threshold:.3f}  "
              f"accuracy {r.accuracy:.3f}")

    chosen = pick_threshold(curve, args.target_accuracy, args.min_coverage)
    if chosen is None:
        best = curve[curve.coverage >= args.min_coverage].accuracy.max()
        sys.exit(f"\nNo threshold reaches {args.target_accuracy:.0%} accuracy at "
                 f"{args.min_coverage:.0%} coverage. The best available is "
                 f"{best:.3f}. Lower --target-accuracy and rerun, and say in the "
                 f"methodology that the target was revised and why.")
    tau = float(chosen.threshold)
    lo, hi = wilson(int(chosen.n_correct), int(chosen.n_retained))
    print(f"\nthreshold {tau:.3f} for a {args.target_accuracy:.0%} target")
    print(f"  retained {int(chosen.n_retained)}/{len(oof)} out-of-fold "
          f"({chosen.coverage:.1%}), accuracy {chosen.accuracy:.3f} "
          f"[{lo:.3f}, {hi:.3f}]")
    below = oof[oof.confidence < tau]
    print(f"  below the threshold: accuracy {below.correct.mean():.3f} on {len(below)}")
    return oof, curve, chosen, tau, lo, hi, below


FUNDER_CANDIDATES = ["funder", "funder_name", "funding_org", "funding_organisation",
                     "lead_funder", "council", "funder_clean"]


def build_verification_sample(labelled, proj, classes, n, reuse=False):
    """Stratified blind coding sheet, plus the key it is scored against."""
    sub = pd.DataFrame({"project_id": coalesce(proj, "project_id"),
                        "abstract_text_clean": coalesce(proj, "abstract_text_clean")})
    funder_col = next((c for c in FUNDER_CANDIDATES if c in proj.columns), None)
    if funder_col:
        sub["funder"] = coalesce(proj, funder_col)
    else:
        print("  no funder column found in projects.csv, leaving that column blank")
    # Drop any funder column already on `labelled`, otherwise the merge makes
    # two columns of the same name and every row value arrives as a Series.
    pool = (labelled[labelled.field_source == "model_assigned"]
            .drop(columns=["funder", "abstract_text_clean"], errors="ignore")
            .merge(sub, on="project_id", how="left"))
    pool = pool[pool.abstract_text_clean.notna() & (pool.abstract_text_clean != "")]

    sheet = ROOT / "data" / "validation" / "discipline_verification_sample.xlsx"
    key_path = sheet.with_name("discipline_verification_KEY.csv")
    if reuse and key_path.exists():
        # Redrawing after a taxonomy change would orphan coding already done,
        # because the stratification is by predicted field. Keep the same
        # projects and refresh only what the model now says about them.
        prev = pd.read_csv(key_path)[["sample_id", "project_id"]]
        sample = prev.merge(pool, on="project_id", how="left")
        missing = int(sample.model_field.isna().sum())
        print(f"  --reuse-sample: kept the {len(prev)} projects already coded"
              + (f", {missing} no longer model-assigned" if missing else ""))
        sample = sample[sample.model_field.notna()].reset_index(drop=True)
    else:
        rng = np.random.RandomState(SEED)
        sample = stratified_sample(pool, min(n, len(pool)), ["tier", "model_field"], rng)
        sample.insert(0, "sample_id", [f"V{i:03d}" for i in range(1, len(sample) + 1)])

    sheet.parent.mkdir(parents=True, exist_ok=True)
    if not reuse:
        write_coding_sheet(sample, classes, sheet)
    key = sample[["sample_id", "project_id", "tier", "model_field",
                  "model_confidence", "model_second_field",
                  "model_second_confidence"]]
    key.to_csv(sheet.with_name("discipline_verification_KEY.csv"), index=False)

    print(f"\nverification sample of {len(sample)} written to {sheet.name}")
    print(sample.groupby("tier").size().to_string())
    print("The sheet carries no model label. Code it first, then score it "
          "against discipline_verification_KEY.csv.")
    return sample


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--crosswalk", choices=["james", "kirsty", "merged10"],
                    default="james",
                    help="which crosswalk variant's gold set and folds to use")
    ap.add_argument("--suffix", default="_mpnet")
    ap.add_argument("--target-accuracy", type=float, default=0.80,
                    help="accuracy the retained tier 2 predictions must reach")
    ap.add_argument("--min-coverage", type=float, default=0.10,
                    help="refuse a threshold that keeps less than this share")
    ap.add_argument("--repeats", type=int, default=1, choices=[1, 2, 3, 4, 5],
                    help="repeats of the 5-fold split used for the curve")
    ap.add_argument("--weight", type=int, default=None,
                    help="CE upweighting factor; omit to select by inner CV")
    ap.add_argument("--sample-size", type=int, default=100)
    ap.add_argument("--proba-only", action="store_true",
                    help="skip the out-of-fold work, reuse the threshold already "
                         "in threshold_summary.json, refit once and write the "
                         "full field probability matrix. About three minutes.")
    ap.add_argument("--reuse-sample", action="store_true",
                    help="keep the projects already in "
                         "discipline_verification_KEY.csv and rewrite only the "
                         "model's answers for them, instead of drawing a fresh "
                         "sample. Use this after changing the taxonomy so the "
                         "coding you have already done stays valid.")
    ap.add_argument("--sample-only", action="store_true",
                    help="regenerate just the verification sheet from an "
                         "existing projects_labelled_final.csv, without "
                         "refitting anything")
    args = ap.parse_args()
    sfx, VARIANT = args.suffix, args.crosswalk
    RESULTS.mkdir(parents=True, exist_ok=True)
    t0 = time.time()

    # --- the frozen evaluation objects, not rebuilt -------------------------
    gold_path, folds_path = DIR / f"gold_{VARIANT}.csv", DIR / f"folds_{VARIANT}.json"
    if not gold_path.exists() or not folds_path.exists():
        sys.exit(f"Missing {gold_path.name} or {folds_path.name}. "
                 f"Run run_variant.py --crosswalk {VARIANT} first.")
    gold = pd.read_csv(gold_path)
    folds = json.load(open(folds_path))["folds"][: args.repeats * 5]
    classes = sorted(gold.primary_field.unique())
    y = gold.primary_field.values
    loc = {p: i for i, p in enumerate(gold.project_id)}
    gold_text = (gold.title_clean.fillna("") + ". "
                 + gold.abstract_text_clean.fillna("")).str.strip().values
    print(f"gold {len(gold)} projects, {len(classes)} classes, "
          f"{len(folds)} folds for the curve")

    if args.sample_only:
        final = DIR / "projects_labelled_final.csv"
        if not final.exists():
            sys.exit(f"--sample-only needs {final.name}, which does not exist yet.")
        print(f"--sample-only: rebuilding the sheet from {final.name}, "
              f"nothing refitted")
        build_verification_sample(pd.read_csv(final),
                                  pd.read_csv(PROJECTS, low_memory=False),
                                  classes, args.sample_size,
                                  reuse=args.reuse_sample)
        return

    # --- features -----------------------------------------------------------
    mapping = load_mapping(VARIANT)
    pemb = np.load(DIR / f"project_embeddings{sfx}.npy")
    pidx = pd.read_csv(DIR / "project_embedding_index.csv")
    prow = {p: i for i, p in enumerate(pidx.project_id)}
    Xg = np.stack([pemb[prow[p]] for p in gold.project_id])

    corpus = pd.read_csv(CORPUS)
    corpus["primary_field"] = corpus.research_subjects.map(lambda s: primary_field(s, mapping))
    cemb = np.load(DIR / f"corpus_embeddings{sfx}.npy")
    cidx = pd.read_csv(DIR / "corpus_embedding_index.csv")
    corpus = corpus.set_index("project_id").reindex(cidx.project_id).reset_index()

    # Same leakage guard as the evaluation: a corpus row that is also a CE
    # project would otherwise train on the thing being labelled.
    ce_ids = set(pd.read_csv(PROJECTS, usecols=["project_id"]).project_id)
    overlap = corpus.project_id.isin(ce_ids).values
    if overlap.any():
        print(f"leakage guard: dropped {int(overlap.sum())} corpus rows that are CE projects")
    ckeep = corpus.primary_field.isin(classes).values & ~overlap
    Xc, yc = cemb[ckeep], corpus.primary_field.values[ckeep]
    ctext = (corpus.title.fillna("") + ". "
             + corpus.abstract_text.fillna("")).str.strip().values[ckeep]
    print(f"corpus in scope {ckeep.sum()}")

    # The projects that need a label.
    proj = pd.read_csv(PROJECTS, low_memory=False)
    # The funder is carried through to the final dataset: RQ3 needs to split
    # discipline by funder, and the missingness that made this classifier
    # necessary is itself a funder effect.
    fcol = next((c for c in FUNDER_CANDIDATES if c in proj.columns), None)
    proj["funder"] = coalesce(proj, fcol) if fcol else pd.NA
    if fcol:
        print(f"funder taken from '{fcol}' ({int(proj.funder.notna().sum())} populated)")
    proj["funder_field"] = proj.research_subjects.map(
        lambda s: primary_field(s, mapping) if pd.notna(s) else None)
    # A handful of funder-labelled projects sit in classes too rare to model
    # (fewer than five gold cases), so run_variant.py dropped them from the
    # evaluation. Their funder label is still the best evidence available, so
    # they stay tier 1 with a field the classifier could never have predicted.
    rare = proj.funder_field.notna() & ~proj.funder_field.isin(classes)
    if rare.any():
        print(f"{int(rare.sum())} funder-labelled projects sit in classes below the "
              f"modelling threshold: {sorted(proj.loc[rare, 'funder_field'].unique())}")
    unlabelled = proj[proj.funder_field.isna()].reset_index(drop=True)
    utext = (unlabelled.title_clean.fillna("") + ". "
             + unlabelled.abstract_text_clean.fillna("")).str.strip().values
    missing_ids = [p for p in unlabelled.project_id if p not in prow]
    if missing_ids:
        sys.exit(f"{len(missing_ids)} projects have no embedding, e.g. {missing_ids[:3]}. "
                 f"Re-run embed_texts_mpnet.py --projects-only first.")
    Xu = np.stack([pemb[prow[p]] for p in unlabelled.project_id])
    print(f"{len(proj)} projects: {int(proj.funder_field.notna().sum())} funder-labelled, "
          f"{len(unlabelled)} to be model-labelled")

    # TF-IDF vocabulary fitted on corpus text only, exactly as in run_variant.
    print("\nfitting TF-IDF on the corpus")
    vec = TfidfVectorizer(sublinear_tf=True, min_df=3, ngram_range=(1, 2),
                          stop_words="english", max_features=200_000)
    Zc = vec.fit_transform(ctext)
    Zg, Zu = vec.transform(gold_text), vec.transform(utext)
    print(f"  {Zc.shape[1]} features ({time.time() - t0:.0f}s)")

    if args.proba_only:
        # Both the threshold and the weight come from the recorded run, so the
        # inner cross-validation is skipped entirely rather than run and then
        # discarded.
        summ = json.load(open(RESULTS / "threshold_summary.json"))
        tau, weight = float(summ["threshold"]), int(summ["weight"])
        oof = chosen = below = lo = hi = None
        print(f"\n--proba-only: reusing threshold {tau:.3f} and weight x{weight} "
              f"from threshold_summary.json, refitting once")
    else:
        weight = args.weight
        if weight is None:
            print("\nselecting the CE upweighting factor by inner cross-validation")
            weight = choose_weight(Xc, yc, Xg, y)
        oof, curve, chosen, tau, lo, hi, below = _oof_and_threshold(
            folds, loc, gold, y, Xc, yc, Zc, Xg, Zg, weight, args, t0)

    # --- 3. final model and labels ------------------------------------------
    print("\nrefitting H on the corpus plus the whole gold set")
    emb, tfidf = fit_h(Xc, yc, Zc, Xg, y, Zg, weight)
    pred, conf, proba = predict_h(emb, tfidf, Xu, Zu)

    # The full probability vector, not just the winner. RQ3 is a distributional
    # question, and soft counting (adding each project's probability to every
    # field) halves the error of hard assignment on the gold set: total
    # variation 0.044 against 0.091, and Engineering's 6.3-point overstatement
    # falls to 0.5. Hard labels stay for anything needing one answer per
    # project; the matrix is what the distribution should be built from.
    known = proj[proj.funder_field.notna()]
    # Columns cover every field with evidence behind it, not just the twelve the
    # model can predict. The funder-labelled projects in rare classes are known
    # facts, and dropping them because the classifier cannot reach those classes
    # would quietly delete real observations from the distribution.
    extra = sorted(set(known.funder_field.dropna()) - set(emb.classes_))
    cols = list(emb.classes_) + extra
    if extra:
        print(f"  including {len(extra)} funder-only fields with no model column: "
              f"{extra}")

    proba_frame = pd.DataFrame(0.0, index=range(len(unlabelled)), columns=cols)
    proba_frame.loc[:, list(emb.classes_)] = proba
    proba_frame.insert(0, "project_id", unlabelled.project_id.values)

    onehot = pd.DataFrame(0.0, index=range(len(known)), columns=cols)
    for i, f in enumerate(known.funder_field.values):
        onehot.iloc[i, onehot.columns.get_loc(f)] = 1.0
    onehot.insert(0, "project_id", known.project_id.values)

    full = pd.concat([onehot, proba_frame], ignore_index=True)
    full.to_csv(DIR / "project_field_probabilities.csv", index=False)
    mass = full[cols].to_numpy().sum()
    print(f"  wrote project_field_probabilities.csv: {len(full)} projects x "
          f"{len(cols)} fields, total mass {mass:.1f} (should equal {len(full)})")

    order = np.argsort(-proba, axis=1)
    unlabelled["model_field"] = pred
    unlabelled["model_confidence"] = conf
    unlabelled["model_second_field"] = emb.classes_[order[:, 1]]
    unlabelled["model_second_confidence"] = proba[np.arange(len(order)), order[:, 1]]

    labelled = proj[["project_id", "title_clean", "funder", "funder_field"]].copy()
    labelled = labelled.merge(
        unlabelled[["project_id", "model_field", "model_confidence",
                    "model_second_field", "model_second_confidence"]],
        on="project_id", how="left")
    labelled["field"] = labelled.funder_field.fillna(labelled.model_field)
    labelled["field_source"] = np.where(labelled.funder_field.notna(),
                                        "funder_assigned", "model_assigned")
    labelled["confidence"] = labelled.model_confidence
    labelled["tier"] = np.select(
        [labelled.funder_field.notna(), labelled.model_confidence >= tau],
        [1, 2], default=3)
    out = DIR / "projects_labelled_final.csv"
    labelled.to_csv(out, index=False)

    print("\n=== tiers ===")
    for tier, n in labelled.tier.value_counts().sort_index().items():
        note = {1: "funder-assigned", 2: "model, at or above threshold",
                3: "model, below threshold"}[tier]
        print(f"  tier {tier}  {n:>5}  {n / len(labelled):>5.1%}  {note}")
    print("\nfield distribution")
    print(labelled.field.value_counts().to_string())

    summary = None if args.proba_only else dict(
        variant=VARIANT, setup="H", weight=int(weight),
        target_accuracy=args.target_accuracy, min_coverage=args.min_coverage,
        threshold=round(tau, 4),
        oof_folds=len(folds), oof_n=int(len(oof)),
        oof_accuracy=round(float(oof.correct.mean()), 4),
        oof_macro_f1=round(float(f1_score(oof.true_field, oof.pred_field,
                                          average="macro")), 4),
        retained_n=int(chosen.n_retained), retained_coverage=round(float(chosen.coverage), 4),
        retained_accuracy=round(float(chosen.accuracy), 4),
        retained_ci=[round(lo, 4), round(hi, 4)],
        below_threshold_accuracy=round(float(below.correct.mean()), 4),
        below_threshold_n=int(len(below)),
        tier_counts={int(k): int(v) for k, v in labelled.tier.value_counts().items()},
        n_projects=int(len(labelled)), n_classes=len(classes))
    if not args.proba_only:
        json.dump(summary, open(RESULTS / "threshold_summary.json", "w"), indent=2)

    # --- 4. blind verification sample ---------------------------------------
    build_verification_sample(labelled, proj, classes, args.sample_size,
                              reuse=args.reuse_sample)
    written = [out.name, "project_field_probabilities.csv"]
    if not args.proba_only:
        written += ["oof_predictions_H.csv", "accuracy_reject_curve.csv",
                    "threshold_summary.json"]
    print(f"\nWrote {', '.join(written)}. Total {time.time() - t0:.0f}s.")


if __name__ == "__main__":
    main()
