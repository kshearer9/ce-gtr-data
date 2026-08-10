"""Does it matter whether the two unlearnable classes are merged before or after training?

Section 3.12 reports a ten-class taxonomy, arrived at by folding Biochemistry,
Genetics and Molecular Biology into Agricultural and Biological Sciences and
Materials Science into Engineering. That was a post-hoc revision, and the
chapter claims that essentially all of the apparent improvement over the
twelve-class figures is definitional rather than the model performing better.

That claim needs a test, because there are two different things the merge could
be doing. It could simply be relabelling the scoreboard, in which case a model
trained on twelve classes and then having its predictions folded down would
score the same as one trained on ten. Or the merge could genuinely help the
model, by giving two starved classes enough examples to learn from, in which
case training on ten would win. The difference matters for how the revision is
described: a definitional change is honest bookkeeping, whereas a real gain
obtained after seeing the results needs declaring as such.

The two arms, both evaluated against the same ten-class ground truth:

    train_ten     gold labelled in the ten-class taxonomy, model trained on
                  ten classes, predictions scored directly.

    fold_after    gold labelled in the twelve-class taxonomy, model trained on
                  twelve classes, predictions then folded to ten and scored
                  against the same ten-class truth.

Both arms run set-up H (the deployed configuration: soft vote of a TF-IDF
classifier and an embedding classifier, both trained on the wider GtR corpus
plus the gold set with gold instances upweighted, the weight chosen by inner
cross-validation) on the SAME twenty-five splits. Using one fold set for both
is what makes the paired test legitimate: the folds stored in
folds_james.json and folds_merged10.json are not the same partitions, because
stratification depends on the label set, so comparing the two existing runs
against each other would not be a paired comparison at all. This script
therefore ignores both stored fold files and rebuilds one partition, stratified
on the ten-class labels, which both arms share.

Everything else is held identical to run_variant.py: the same seed, the same
weight grid, the same regularisation, the same leakage guard, the same scoring
function. The only thing that varies between arms is when the merge happens.

Run from the repo root, about an hour:

    /opt/anaconda3/bin/python scripts/classification/test_merge_timing.py

Writes results/taxonomy_merge_timing.csv, twenty-five rows per arm, and prints
the paired comparison. Whatever it prints is the figure to quote; if it differs
from the p-values currently in the draft, the draft is wrong and this is right,
because this is the version with an artefact behind it.
"""
from pathlib import Path
import re
import sys
import time

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.sparse import vstack as spvstack
from scipy.stats import wilcoxon
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, f1_score
from sklearn.model_selection import RepeatedStratifiedKFold, StratifiedKFold

ROOT = Path(__file__).resolve().parents[2]
DIR = ROOT / "data" / "classification"
RESULTS = DIR / "results"
XW = ROOT / "data" / "crosswalk" / "crosswalk_gtr_to_openalex_FINAL.xlsx"
PROJECTS = ROOT / "data" / "cleaned" / "merged" / "projects.csv"
CORPUS = DIR / "gtr_tagged_corpus.csv"
SFX = "_mpnet"

# Identical to run_variant.py. Changing any of these invalidates the comparison
# with the set-up H figures reported in the chapter.
SEED = 20260803
C_TFIDF, C_EMB, MAX_ITER = 1.0, 3.0, 3000
WEIGHT_GRID = [1, 5, 10, 25, 50, 100]
MIN_CLASS = 5

# Subject-level merge, applied when building the ten-class gold set.
MERGED10 = {
    "Biomolecules & biochemistry": "Agricultural and Biological Sciences",
    "Omic sciences & technologies": "Agricultural and Biological Sciences",
    "Cell biology": "Agricultural and Biological Sciences",
    "Genetics & development": "Agricultural and Biological Sciences",
    "Materials sciences": "Engineering",
}

# Field-level merge, applied to twelve-class predictions to fold them down.
FIELD_MERGE = {
    "Biochemistry, Genetics and Molecular Biology": "Agricultural and Biological Sciences",
    "Materials Science": "Engineering",
}

EXTENSION = {
    "Medical & health interface": "Medicine",
    "Pol. sci. & internat. studies": "Social Sciences",
    "Atmospheric phys. & chemistry": "Earth and Planetary Sciences",
    "Social Policy": "Social Sciences",
    "Particle physics - experiment": "Physics and Astronomy",
    "Astronomy - observation": "Physics and Astronomy",
    "Optics, photonics & lasers": "Physics and Astronomy",
    "Supercond, magn. &quant.fluids": "Physics and Astronomy",
    "Linguistics": "Arts and Humanities",
    "Archaeology": "Arts and Humanities",
    "Particle astrophysics": "Physics and Astronomy",
    "Particle physics - theory": "Physics and Astronomy",
    "Demography & human geography": "Social Sciences",
    "Philosophy": "Arts and Humanities",
    "Planetary science": "Earth and Planetary Sciences",
    "Music": "Arts and Humanities",
    "Drama & theatre studies": "Arts and Humanities",
    "Area Studies": "Social Sciences",
    "Astronomy - theory": "Physics and Astronomy",
    "Solar & terrestrial physics": "Physics and Astronomy",
    "Nuclear physics": "Physics and Astronomy",
    "Plasma physics": "Physics and Astronomy",
    "Social Work": "Social Sciences",
    "Library & information studies": "Social Sciences",
    "Demography": "Social Sciences",
    "Classics": "Arts and Humanities",
    "Facility Development": None,
    "Dance": "Arts and Humanities",
}

PATTERN = re.compile(r"^\s*(.+?)\s*\((\d+(?:\.\d+)?)%\)\s*$")


def parse_subjects(value):
    out = []
    for part in str(value).split(";"):
        part = part.strip()
        if not part:
            continue
        m = PATTERN.match(part)
        out.append((m.group(1).strip(), float(m.group(2))) if m else (part, 100.0))
    return out


def load_mapping(merge_subjects):
    """The crosswalk, optionally with the subject-level merge applied."""
    ws = load_workbook(XW)["Crosswalk (FINAL)"]
    mapping = {}
    for row in range(7, ws.max_row + 1):
        subject = ws.cell(row, 1).value
        if subject:
            field = ws.cell(row, 3).value
            mapping[subject] = None if field == "EXCLUDED" else field
    mapping.update(EXTENSION)
    if merge_subjects:
        mapping.update(MERGED10)
    return mapping


def primary_field(value, mapping):
    mapped = [(mapping[s], w) for s, w in parse_subjects(value) if mapping.get(s)]
    return max(mapped, key=lambda t: t[1])[0] if mapped else None


def build_gold(mapping, label):
    projects = pd.read_csv(PROJECTS)
    gold = projects[projects.research_subjects.notna()].copy()
    gold["primary_field"] = gold.research_subjects.map(lambda s: primary_field(s, mapping))
    gold = gold[gold.primary_field.notna()].reset_index(drop=True)
    counts = gold.primary_field.value_counts()
    rare = counts[counts < MIN_CLASS].index.tolist()
    kept = gold[~gold.primary_field.isin(rare)].reset_index(drop=True)
    print(f"  {label}: {len(kept)} projects, {kept.primary_field.nunique()} classes")
    return kept


def score(y_true, pred):
    """The same scoring used everywhere else, so the numbers are comparable."""
    return dict(macro_f1=f1_score(y_true, pred, average="macro"),
                accuracy=accuracy_score(y_true, pred))


def prepare(gold, mapping, sfx=SFX):
    """Embeddings, text and corpus for one labelling of the gold set."""
    classes = sorted(gold.primary_field.unique())
    text = (gold.title_clean.fillna("") + ". "
            + gold.abstract_text_clean.fillna("")).str.strip().values

    pemb = np.load(DIR / f"project_embeddings{sfx}.npy")
    pidx = pd.read_csv(DIR / "project_embedding_index.csv")
    if len(pidx) != len(pemb):
        sys.exit(f"project_embedding_index.csv has {len(pidx)} rows against "
                 f"{len(pemb)} embeddings. Re-run embed_texts_mpnet.py --projects-only.")
    prow = {p: i for i, p in enumerate(pidx.project_id)}
    X = np.stack([pemb[prow[p]] for p in gold.project_id])

    corpus = pd.read_csv(CORPUS)
    corpus["primary_field"] = corpus.research_subjects.map(
        lambda s: primary_field(s, mapping))
    cemb = np.load(DIR / f"corpus_embeddings{sfx}.npy")
    cidx = pd.read_csv(DIR / "corpus_embedding_index.csv")
    if len(cidx) != len(cemb):
        sys.exit(f"corpus_embedding_index.csv has {len(cidx)} rows against "
                 f"{len(cemb)} embeddings. Re-run embed_texts_mpnet.py.")
    corpus = corpus.set_index("project_id").reindex(cidx.project_id).reset_index()

    # Same leakage guard as the main run: a corpus row that is itself a CE
    # project would otherwise be in training and evaluation at once.
    ce_ids = set(pd.read_csv(PROJECTS, usecols=["project_id"]).project_id)
    overlap = corpus.project_id.isin(ce_ids)
    ckeep = corpus.primary_field.isin(classes).values & ~overlap.values
    print(f"    leakage guard dropped {int(overlap.sum())} corpus rows; "
          f"corpus in scope {int(ckeep.sum())}")
    Xc, yc = cemb[ckeep], corpus.primary_field.values[ckeep]
    ctext = (corpus.title.fillna("") + ". "
             + corpus.abstract_text.fillna("")).str.strip().values[ckeep]
    return dict(classes=classes, y=gold.primary_field.values, X=X, text=text,
                Xc=Xc, yc=yc, ctext=ctext,
                loc={p: i for i, p in enumerate(gold.project_id)})


def fit_emb(Xtr, ytr, w=None):
    clf = LogisticRegression(max_iter=MAX_ITER, C=C_EMB, class_weight="balanced")
    clf.fit(Xtr, ytr, sample_weight=w)
    return clf


def run_arm(arm, D, folds, y_truth10, fold_predictions):
    """Set-up H over the shared folds. Returns one row per split."""
    print(f"\n[{arm}] fitting TF-IDF on corpus text once")
    t0 = time.time()
    vec = TfidfVectorizer(sublinear_tf=True, min_df=3, ngram_range=(1, 2),
                          stop_words="english", max_features=200_000)
    Zc = vec.fit_transform(D["ctext"])
    Zgold = vec.transform(D["text"])
    print(f"[{arm}] {Zc.shape[1]} features ({time.time() - t0:.0f}s)")

    X, y, Xc, yc, loc = D["X"], D["y"], D["Xc"], D["yc"], D["loc"]
    rows = []
    for fold in folds:
        t1 = time.time()
        tr = [loc[p] for p in fold["train"]]
        te = [loc[p] for p in fold["test"]]

        # Weight chosen on inner folds of the training portion only, never on
        # the evaluation fold (Cawley and Talbot, 2010).
        inner = StratifiedKFold(n_splits=3, shuffle=True, random_state=7)
        best, best_score = WEIGHT_GRID[0], -1.0
        for mult in WEIGHT_GRID:
            inner_scores = []
            for itr, ite in inner.split(X[tr], y[tr]):
                sub, held = np.asarray(tr)[itr], np.asarray(tr)[ite]
                w = np.concatenate([np.ones(len(yc)), np.full(len(sub), float(mult))])
                clf = fit_emb(np.vstack([Xc, X[sub]]), np.concatenate([yc, y[sub]]), w)
                inner_scores.append(
                    f1_score(y[held], clf.predict(X[held]), average="macro"))
            if np.mean(inner_scores) > best_score:
                best, best_score = mult, float(np.mean(inner_scores))

        w = np.concatenate([np.ones(len(yc)), np.full(len(tr), float(best))])
        ytr_g = np.concatenate([yc, y[tr]])
        g_model = fit_emb(np.vstack([Xc, X[tr]]), ytr_g, w)

        tfidf_h = LogisticRegression(max_iter=MAX_ITER, C=C_TFIDF,
                                     class_weight="balanced")
        tfidf_h.fit(spvstack([Zc, Zgold[tr]]), ytr_g, sample_weight=w)
        assert list(tfidf_h.classes_) == list(g_model.classes_)
        vote = (tfidf_h.predict_proba(Zgold[te]) + g_model.predict_proba(X[te])) / 2
        pred = g_model.classes_[np.argmax(vote, axis=1)]

        # This is the whole point of the script: the fold_after arm predicts in
        # the twelve-class space and is collapsed to ten here, AFTER training,
        # then scored against the same ten-class truth as the other arm.
        pred10 = pd.Series(pred).replace(FIELD_MERGE).values
        truth10 = y_truth10[te]

        rows.append(dict(arm=arm, split=fold["split"], weight=best,
                         **score(truth10, pred10)))
        fold_predictions.append(pd.DataFrame(
            dict(arm=arm, split=fold["split"], project_id=fold["test"],
                 truth10=truth10, pred_raw=pred, pred10=pred10)))
        print(f"  split {fold['split']:2d}  macro-F1 {rows[-1]['macro_f1']:.3f}  "
              f"acc {rows[-1]['accuracy']:.3f}  (weight x{best}, "
              f"{time.time() - t1:.0f}s)", flush=True)
    return rows


def main() -> None:
    t_start = time.time()
    for path in (XW, PROJECTS, CORPUS):
        if not path.exists():
            sys.exit(f"Missing {path}.")
    RESULTS.mkdir(parents=True, exist_ok=True)

    print("Building both labellings of the gold set")
    map10 = load_mapping(merge_subjects=True)
    map12 = load_mapping(merge_subjects=False)
    gold10 = build_gold(map10, "ten-class")
    gold12 = build_gold(map12, "twelve-class")

    # Both arms must score the same projects in the same order, or the paired
    # test is comparing different things. Refuse rather than silently realign.
    if list(gold10.project_id) != list(gold12.project_id):
        sys.exit("The two labellings cover different projects, so the arms are "
                 "not comparable. Check MIN_CLASS and the rare-class drop.")

    # One partition, stratified on the ten-class labels, used by both arms.
    # The stored fold files differ between variants because stratification
    # depends on the label set, so neither is reused here.
    rskf = RepeatedStratifiedKFold(n_splits=5, n_repeats=5, random_state=SEED)
    folds = [dict(split=i, train=gold10.project_id.iloc[tr].tolist(),
                  test=gold10.project_id.iloc[te].tolist())
             for i, (tr, te) in enumerate(
                 rskf.split(np.zeros(len(gold10)), gold10.primary_field))]
    print(f"  {len(folds)} shared splits, stratified on the ten-class labels")

    print("\nPreparing the ten-class arm")
    D10 = prepare(gold10, map10)
    print("Preparing the twelve-class arm")
    D12 = prepare(gold12, map12)

    # The ten-class truth, indexed the same way for both arms.
    y_truth10 = gold10.primary_field.values

    preds = []
    rows = run_arm("train_ten", D10, folds, y_truth10, preds)
    rows += run_arm("fold_after", D12, folds, y_truth10, preds)

    res = pd.DataFrame(rows)
    res.to_csv(RESULTS / "taxonomy_merge_timing.csv", index=False)
    pd.concat(preds).to_csv(RESULTS / "taxonomy_merge_timing_predictions.csv",
                            index=False)

    a = res[res.arm == "train_ten"].sort_values("split").reset_index(drop=True)
    b = res[res.arm == "fold_after"].sort_values("split").reset_index(drop=True)

    print("\n" + "=" * 70)
    print("MERGE BEFORE OR AFTER TRAINING, ON IDENTICAL SPLITS")
    print("=" * 70)
    print(res.groupby("arm")[["macro_f1", "accuracy"]]
          .agg(["mean", "std"]).round(3).to_string())

    for metric in ("macro_f1", "accuracy"):
        d = a[metric] - b[metric]
        wins = int((a[metric] > b[metric]).sum())
        ties = int((a[metric] == b[metric]).sum())
        if d.abs().sum() == 0:
            print(f"\n{metric}: the two arms are identical on every split.")
            continue
        p = wilcoxon(a[metric], b[metric]).pvalue
        print(f"\n{metric}: train_ten minus fold_after = {d.mean():+.4f} "
              f"(sd {d.std():.4f})")
        print(f"  train_ten wins {wins} of {len(a)}, ties {ties}, "
              f"loses {len(a) - wins - ties}")
        print(f"  Wilcoxon signed-rank p = {p:.3f}")

    print("\nReading: a difference indistinguishable from zero means the merge is\n"
          "definitional, the model is not learning anything extra from having ten\n"
          "classes to train on, and the ten-class figures may be reported as a\n"
          "relabelling of the twelve-class ones. A clear win for train_ten would\n"
          "mean the merge genuinely helps the model, which is a post-hoc gain and\n"
          "must be described as one.")
    print(f"\nTotal {time.time() - t_start:.0f}s. Wrote "
          f"taxonomy_merge_timing.csv and taxonomy_merge_timing_predictions.csv")


if __name__ == "__main__":
    main()
