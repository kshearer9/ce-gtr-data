"""Full classification comparison for one crosswalk variant.

Rebuilds the gold standard and cross-validation folds from a chosen crosswalk,
re-maps the wider GtR corpus with the same crosswalk, then runs the method
bake-off and the training-corpus set-ups A to H. Every output is tagged with the
variant name so runs never overwrite each other.

Crosswalk variants
    james   the verified crosswalk, ASJC placement plus corpus evidence
    kirsty  as above with the second reviewer's six amendments applied

Set-ups
    A  CE project abstracts only
    B  publication abstracts only            (leakage-guarded)
    C  projects + publications               (leakage-guarded)
    D  wider GtR corpus only
    E  corpus + projects, unweighted
    F  corpus subsampled to the CE class distribution
    G  corpus + projects, CE upweighted      (Jiang and Zhai, 2007)
    H  as G, but a soft-vote of TF-IDF and embedding classifiers

G and H select the weighting factor by nested cross-validation, so the reported
figure is not inflated by choosing on the evaluation folds (Cawley and Talbot,
2010).

CAUTION: the two variants yield different numbers of classes, because the
kirsty variant dissolves Chemical Engineering. Macro-F1 across different class
counts is not directly comparable; the class count is printed for this reason.

Run from the repo root, e.g.
    /opt/anaconda3/bin/python scripts/classification/run_variant.py --crosswalk kirsty
    /opt/anaconda3/bin/python scripts/classification/run_variant.py --crosswalk james --skip-h
"""
from pathlib import Path
import argparse
import json
import re
import sys
import time

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, f1_score
from sklearn.model_selection import RepeatedStratifiedKFold, StratifiedKFold
from sklearn.naive_bayes import MultinomialNB
from sklearn.pipeline import make_pipeline

ROOT = Path(__file__).resolve().parents[2]
DIR = ROOT / "data" / "classification"
RESULTS = DIR / "results"
XW = ROOT / "data" / "crosswalk" / "crosswalk_gtr_to_openalex_FINAL.xlsx"
PROJECTS = ROOT / "data" / "cleaned" / "merged" / "projects.csv"
CORPUS = DIR / "gtr_tagged_corpus.csv"

SEED = 20260803
C_TFIDF, C_EMB, MAX_ITER = 1.0, 3.0, 3000
WEIGHT_GRID = [1, 5, 10, 25, 50, 100]
MIN_CLASS = 5

# Second reviewer's amendments (see crosswalk_kirsty_verified.xlsx)
KIRSTY = {
    "Bioengineering": "Engineering",
    "Materials processing": "Materials Science",
    "Catalysis & surfaces": "Chemistry",
    "Environmental engineering": "Engineering",
    "Process engineering": "Engineering",
    "Environmental Planning": "Environmental Science",
}

# Corpus-only subjects, training use only (never appear in the 1,380 CE projects)
EXTENSION = {
    "Medical & health interface": "Medicine",
    "Pol. sci. & internat. studies": "Social Sciences",
    "Atmospheric phys. & chemistry": "Earth and Planetary Sciences",
    "Social Policy": "Social Sciences",
    "Particle physics - experiment": "Physics and Astronomy",
    "Astronomy - observation": "Physics and Astronomy",
    "Optics, photonics & lasers": "Physics and Astronomy",
    "Supercond, magn. &quant.fluids": "Physics and Astronomy",
    "Linguistics": "Arts and Humanities",
    "Archaeology": "Arts and Humanities",
    "Particle astrophysics": "Physics and Astronomy",
    "Particle physics - theory": "Physics and Astronomy",
    "Demography & human geography": "Social Sciences",
    "Philosophy": "Arts and Humanities",
    "Planetary science": "Earth and Planetary Sciences",
    "Music": "Arts and Humanities",
    "Drama & theatre studies": "Arts and Humanities",
    "Area Studies": "Social Sciences",
    "Astronomy - theory": "Physics and Astronomy",
    "Solar & terrestrial physics": "Physics and Astronomy",
    "Nuclear physics": "Physics and Astronomy",
    "Plasma physics": "Physics and Astronomy",
    "Social Work": "Social Sciences",
    "Library & information studies": "Social Sciences",
    "Demography": "Social Sciences",
    "Classics": "Arts and Humanities",
    "Facility Development": None,
    "Dance": "Arts and Humanities",
}

PATTERN = re.compile(r"^\s*(.+?)\s*\((\d+(?:\.\d+)?)%\)\s*$")


def parse_subjects(value):
    out = []
    for part in str(value).split(";"):
        part = part.strip()
        if not part:
            continue
        m = PATTERN.match(part)
        out.append((m.group(1).strip(), float(m.group(2))) if m else (part, 100.0))
    return out


def load_mapping(variant):
    ws = load_workbook(XW)["Crosswalk (FINAL)"]
    mapping = {}
    for row in range(7, ws.max_row + 1):
        subject = ws.cell(row, 1).value
        if subject:
            field = ws.cell(row, 3).value
            mapping[subject] = None if field == "EXCLUDED" else field
    mapping.update(EXTENSION)
    if variant == "kirsty":
        mapping.update(KIRSTY)
    return mapping


def primary_field(value, mapping):
    mapped = [(mapping[s], w) for s, w in parse_subjects(value) if mapping.get(s)]
    return max(mapped, key=lambda t: t[1])[0] if mapped else None


def build_gold(variant, mapping):
    projects = pd.read_csv(PROJECTS)
    gold = projects[projects.research_subjects.notna()].copy()
    gold["primary_field"] = gold.research_subjects.map(lambda s: primary_field(s, mapping))
    gold = gold[gold.primary_field.notna()].reset_index(drop=True)

    counts = gold.primary_field.value_counts()
    rare = counts[counts < MIN_CLASS].index.tolist()
    kept = gold[~gold.primary_field.isin(rare)].reset_index(drop=True)
    print(f"[{variant}] gold {len(kept)} projects, {kept.primary_field.nunique()} classes "
          f"(dropped {len(rare)} rare classes covering {len(gold) - len(kept)} projects)")
    print(kept.primary_field.value_counts().to_string())

    rskf = RepeatedStratifiedKFold(n_splits=5, n_repeats=5, random_state=SEED)
    folds = [dict(split=i, train=kept.project_id.iloc[tr].tolist(),
                  test=kept.project_id.iloc[te].tolist())
             for i, (tr, te) in enumerate(rskf.split(np.zeros(len(kept)), kept.primary_field))]
    kept.to_csv(DIR / f"gold_{variant}.csv", index=False)
    json.dump({"seed": SEED, "folds": folds}, open(DIR / f"folds_{variant}.json", "w"))
    return kept, folds


def score(y_true, pred):
    return dict(macro_f1=f1_score(y_true, pred, average="macro"),
                accuracy=accuracy_score(y_true, pred))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--crosswalk", choices=["james", "kirsty"], required=True)
    ap.add_argument("--suffix", default="_mpnet", help="embedding suffix")
    ap.add_argument("--skip-h", action="store_true", help="skip the slow ensemble set-up")
    args = ap.parse_args()
    variant, sfx = args.crosswalk, args.suffix
    RESULTS.mkdir(parents=True, exist_ok=True)
    t_start = time.time()

    mapping = load_mapping(variant)
    gold, folds = build_gold(variant, mapping)
    classes = sorted(gold.primary_field.unique())
    y = gold.primary_field.values
    loc = {p: i for i, p in enumerate(gold.project_id)}
    gold_text = (gold.title_clean.fillna("") + ". "
                 + gold.abstract_text_clean.fillna("")).str.strip().values

    pemb = np.load(DIR / f"project_embeddings{sfx}.npy")
    pidx = pd.read_csv(DIR / "project_embedding_index.csv")
    prow = {p: i for i, p in enumerate(pidx.project_id)}
    X = np.stack([pemb[prow[p]] for p in gold.project_id])

    bemb = np.load(DIR / f"publication_embeddings{sfx}.npy")
    bidx = pd.read_csv(DIR / "publication_embedding_index.csv")
    bkeep = bidx.field.isin(classes).values
    Xb, yb, bproj = bemb[bkeep], bidx.field.values[bkeep], bidx.project_id.values[bkeep]

    # corpus: re-map labels with this variant's crosswalk, keep rows in scope
    corpus = pd.read_csv(CORPUS)
    corpus["primary_field"] = corpus.research_subjects.map(lambda s: primary_field(s, mapping))
    cemb = np.load(DIR / f"corpus_embeddings{sfx}.npy")
    cidx = pd.read_csv(DIR / "corpus_embedding_index.csv")
    corpus = corpus.set_index("project_id").reindex(cidx.project_id).reset_index()

    # LEAKAGE GUARD. The corpus was collected with the CE projects known at the
    # time excluded. Any project added to the CE set since then is still sitting
    # in it, and would otherwise appear in the training data and the evaluation
    # set at once. Exclude by project id against the current CE set, so this
    # holds however the CE set changes.
    ce_ids = set(pd.read_csv(PROJECTS, usecols=["project_id"]).project_id)
    overlap = corpus.project_id.isin(ce_ids)
    if overlap.any():
        print(f"  leakage guard: dropped {int(overlap.sum())} corpus rows that are "
              f"CE projects", flush=True)
    ckeep = corpus.primary_field.isin(classes).values & ~overlap.values
    Xc, yc = cemb[ckeep], corpus.primary_field.values[ckeep]
    ctext = (corpus.title.fillna("") + ". " + corpus.abstract_text.fillna("")).str.strip().values[ckeep]
    print(f"[{variant}] publications in scope {bkeep.sum()} | corpus in scope {ckeep.sum()}\n")

    # Set-up H needs TF-IDF features over the corpus. The vocabulary is fitted
    # ONCE on corpus text only, which is disjoint from every CE project, so no
    # test-fold text informs the vectoriser and there is no leakage. Refitting
    # it per fold would cost ~25x more for no methodological gain.
    tfidf_h_vec = None
    if not args.skip_h:
        t0 = time.time()
        tfidf_h_vec = TfidfVectorizer(sublinear_tf=True, min_df=3, ngram_range=(1, 2),
                                      stop_words="english", max_features=200_000)
        Zc = tfidf_h_vec.fit_transform(ctext)
        Zgold = tfidf_h_vec.transform(gold_text)
        print(f"[{variant}] TF-IDF fitted on corpus once: {Zc.shape[1]} features "
              f"({time.time() - t0:.0f}s)")

    rng = np.random.default_rng(SEED)
    target = gold.primary_field.value_counts(normalize=True)
    matched = np.concatenate([
        rng.choice(np.flatnonzero(yc == cls),
                   size=min(int((yc == cls).sum()), max(1, int(round(prop * len(yc))))),
                   replace=False)
        for cls, prop in target.items() if (yc == cls).sum() > 0])

    def fit_emb(Xtr, ytr, w=None):
        clf = LogisticRegression(max_iter=MAX_ITER, C=C_EMB, class_weight="balanced")
        clf.fit(Xtr, ytr, sample_weight=w)
        return clf

    # ---------------- method bake-off ----------------
    print(f"[{variant}] method bake-off")
    bake = []
    for fold in folds:
        tr = [loc[p] for p in fold["train"]]
        te = [loc[p] for p in fold["test"]]
        tfidf = make_pipeline(TfidfVectorizer(sublinear_tf=True, min_df=2,
                                              ngram_range=(1, 2), stop_words="english"),
                              LogisticRegression(max_iter=MAX_ITER, C=C_TFIDF,
                                                 class_weight="balanced"))
        tfidf.fit(gold_text[tr], y[tr])
        nb = make_pipeline(TfidfVectorizer(sublinear_tf=True, min_df=2,
                                           ngram_range=(1, 2), stop_words="english"),
                           MultinomialNB()).fit(gold_text[tr], y[tr])
        emb = fit_emb(X[tr], y[tr])
        p_tfidf, p_emb = tfidf.predict_proba(gold_text[te]), emb.predict_proba(X[te])
        vote = (p_tfidf + p_emb) / 2
        cent = np.stack([X[tr][y[tr] == c].mean(axis=0) for c in np.unique(y[tr])])
        cent /= np.linalg.norm(cent, axis=1, keepdims=True)
        for name, pred in [
            ("M1_tfidf_lr", tfidf.classes_[np.argmax(p_tfidf, axis=1)]),
            ("M1b_tfidf_nb", nb.predict(gold_text[te])),
            ("M2_emb_lr", emb.classes_[np.argmax(p_emb, axis=1)]),
            ("M3_centroid", np.unique(y[tr])[np.argmax(X[te] @ cent.T, axis=1)]),
            ("M6_softvote", tfidf.classes_[np.argmax(vote, axis=1)]),
        ]:
            bake.append(dict(method=name, split=fold["split"], **score(y[te], pred)))
    bake = pd.DataFrame(bake)
    bake.to_csv(RESULTS / f"bakeoff_{variant}.csv", index=False)
    print(bake.groupby("method")[["macro_f1", "accuracy"]].agg(["mean", "std"]).round(3).to_string())

    # ---------------- corpus set-ups ----------------
    print(f"\n[{variant}] corpus set-ups")
    rows, chosen = [], []
    for fold in folds:
        tr = [loc[p] for p in fold["train"]]
        te = [loc[p] for p in fold["test"]]
        guard = ~np.isin(bproj, fold["test"])
        for name, (Xtr, ytr) in {
            "A": (X[tr], y[tr]),
            "B": (Xb[guard], yb[guard]),
            "C": (np.vstack([X[tr], Xb[guard]]), np.concatenate([y[tr], yb[guard]])),
            "D": (Xc, yc),
            "E": (np.vstack([Xc, X[tr]]), np.concatenate([yc, y[tr]])),
            "F": (Xc[matched], yc[matched]),
        }.items():
            rows.append(dict(setup=name, split=fold["split"],
                             **score(y[te], fit_emb(Xtr, ytr).predict(X[te]))))

        inner = StratifiedKFold(n_splits=3, shuffle=True, random_state=7)
        best, best_score = WEIGHT_GRID[0], -1.0
        for mult in WEIGHT_GRID:
            inner_scores = []
            for itr, ite in inner.split(X[tr], y[tr]):
                sub, held = np.asarray(tr)[itr], np.asarray(tr)[ite]
                w = np.concatenate([np.ones(len(yc)), np.full(len(sub), float(mult))])
                clf = fit_emb(np.vstack([Xc, X[sub]]), np.concatenate([yc, y[sub]]), w)
                inner_scores.append(f1_score(y[held], clf.predict(X[held]), average="macro"))
            if np.mean(inner_scores) > best_score:
                best, best_score = mult, float(np.mean(inner_scores))
        chosen.append(best)
        w = np.concatenate([np.ones(len(yc)), np.full(len(tr), float(best))])
        Xtr_g, ytr_g = np.vstack([Xc, X[tr]]), np.concatenate([yc, y[tr]])
        g_model = fit_emb(Xtr_g, ytr_g, w)
        rows.append(dict(setup="G", split=fold["split"],
                         **score(y[te], g_model.predict(X[te]))))

        if not args.skip_h:
            from scipy.sparse import vstack as spvstack
            Ztr = spvstack([Zc, Zgold[tr]])
            tfidf_h = LogisticRegression(max_iter=MAX_ITER, C=C_TFIDF,
                                         class_weight="balanced")
            tfidf_h.fit(Ztr, ytr_g, sample_weight=w)
            assert list(tfidf_h.classes_) == list(g_model.classes_)
            vote = (tfidf_h.predict_proba(Zgold[te]) + g_model.predict_proba(X[te])) / 2
            rows.append(dict(setup="H", split=fold["split"],
                             **score(y[te], g_model.classes_[np.argmax(vote, axis=1)])))
        print(f"  split {fold['split']:2d} done (G weight x{best}) "
              f"[{time.time() - t_start:.0f}s]", flush=True)

    res = pd.DataFrame(rows)
    res.to_csv(RESULTS / f"setups_{variant}.csv", index=False)
    summary = res.groupby("setup")[["macro_f1", "accuracy"]].agg(["mean", "std"]).round(3)
    summary.to_csv(RESULTS / f"setups_summary_{variant}.csv")
    print(f"\n=== [{variant}] training-corpus comparison "
          f"({len(classes)} classes, {len(gold)} projects) ===")
    print(summary.to_string())
    print(f"\nG weights chosen: {pd.Series(chosen).value_counts().to_dict()}")
    print(f"Total {time.time() - t_start:.0f}s. Results in {RESULTS}")


if __name__ == "__main__":
    main()
