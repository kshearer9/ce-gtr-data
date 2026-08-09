"""Evaluate the classifier against the multi-label truth already in the data.

Strict accuracy asks whether the model picked the same single field as the gold
standard. That gold standard is itself a collapse: GtR tags each project with
weighted research subjects, and the pipeline keeps only the highest-weighted
one. Of the 287 gold projects, 200 carry more than one subject and 161 span
more than one OpenAlex field once the crosswalk is applied. So for over half the
evaluation set, strict accuracy is asking the model to reproduce an argmax over
several defensible answers, and scoring the near-misses as though they were
nonsense.

This script reports the classifier against the multi-label truth the funders
actually supplied. It changes no predictions and retrains nothing. It is a
measurement correction, not an improvement, and it should be described that way.

Four figures come out, in ascending permissiveness:

    strict          model's top-1 equals the funder's highest-weighted field.
                    The conservative headline. Lead with this.

    partial credit  each prediction scored by the weight the funder placed on
                    whichever field the model chose, against the ceiling of
                    always choosing their top field. Reported as a proportion
                    of achievable credit. This is the honest headline for an
                    interdisciplinary corpus: it neither ignores second choices
                    nor treats a 10% field as equal to a 70% one.

    rank profile    where the model's choice sits in the funder's own ordering.
                    Not a score, a diagnostic, and the most informative single
                    table here.

    any tagged      model's top-1 is any field the funder tagged at all.
                    Report it, do not lead with it. It gives full marks for
                    naming a field the funder weighted at 10%.

A fifth, primary-or-crosswalk-secondary, is included because the secondary
column was filled in when the crosswalk was built and predates every result,
which makes it a stronger relaxation than anything invented afterwards.

Nothing here is a substitute for reporting strict accuracy. The case for the
looser figures is substantive, not cosmetic: the object of study is
interdisciplinary research, a single forced label is a poor instrument for it,
and the gap between strict and relaxed is itself evidence of that. Make that
argument explicitly or the numbers will read as inflation.

    /opt/anaconda3/bin/python scripts/classification/evaluate_multilabel.py

Seconds to run. Writes results/multilabel_evaluation.csv.
"""
from pathlib import Path
import argparse
import re
import sys

import numpy as np
import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DIR = ROOT / "data" / "classification"
RESULTS = DIR / "results"
OOF = RESULTS / "oof_predictions_H.csv"
PROJECTS = ROOT / "data" / "cleaned" / "merged" / "projects.csv"
XW = ROOT / "data" / "crosswalk" / "crosswalk_gtr_to_openalex_FINAL.xlsx"

# The merge applied on the project side, so the funder's fields land in the same
# ten-class space as the predictions being scored.
FIELD_MERGE = {
    "Biochemistry, Genetics and Molecular Biology": "Agricultural and Biological Sciences",
    "Materials Science": "Engineering",
}

PATTERN = re.compile(r"^\s*(.+?)\s*\((\d+(?:\.\d+)?)%\)\s*$")


def parse_subjects(value):
    """GtR research subjects arrive as 'Subject (60%); Other (40%)'."""
    out = []
    for part in str(value).split(";"):
        part = part.strip()
        if not part:
            continue
        m = PATTERN.match(part)
        out.append((m.group(1).strip(), float(m.group(2))) if m else (part, 100.0))
    return out


def load_crosswalk(merge=True):
    ws = load_workbook(XW)["Crosswalk (FINAL)"]
    primary, secondary = {}, {}
    for row in range(7, ws.max_row + 1):
        subject = ws.cell(row, 1).value
        if not subject:
            continue
        p, s = ws.cell(row, 3).value, ws.cell(row, 4).value
        primary[subject] = None if p == "EXCLUDED" else p
        secondary[subject] = s
    if merge:
        primary = {k: FIELD_MERGE.get(v, v) for k, v in primary.items()}
        secondary = {k: FIELD_MERGE.get(v, v) for k, v in secondary.items()}
    return primary, secondary


def field_weights(value, primary):
    """The funder's own weighted distribution over fields for one project.

    Two subjects mapping to the same field have their weights added, which is
    why a project can carry four subjects and only two fields.
    """
    weights = {}
    for subject, pct in parse_subjects(value):
        field = primary.get(subject)
        if field:
            weights[field] = weights.get(field, 0.0) + pct
    total = sum(weights.values())
    return {k: v / total for k, v in weights.items()} if total else {}


def allowed_with_secondary(value, primary, secondary):
    """Primary field of the top subject, plus that subject's declared secondary."""
    mapped = [(s, w) for s, w in parse_subjects(value) if primary.get(s)]
    if not mapped:
        return set()
    top = max(mapped, key=lambda t: t[1])[0]
    out = {primary[top]}
    if secondary.get(top):
        out.add(secondary[top])
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-merge", action="store_true",
                    help="score against the unmerged twelve-class fields")
    args = ap.parse_args()

    for path in (OOF, PROJECTS, XW):
        if not path.exists():
            sys.exit(f"Missing {path}. Run apply_classifier.py first.")

    primary, secondary = load_crosswalk(merge=not args.no_merge)
    oof = pd.read_csv(OOF)
    proj = pd.read_csv(PROJECTS, low_memory=False)
    subjects = dict(zip(proj.project_id, proj.research_subjects))

    missing = [p for p in oof.project_id if p not in subjects]
    if missing:
        sys.exit(f"{len(missing)} evaluated projects are not in projects.csv, "
                 f"e.g. {missing[:3]}. The two files are out of step.")

    oof["weights"] = oof.project_id.map(lambda p: field_weights(subjects[p], primary))
    oof["n_subjects"] = oof.project_id.map(lambda p: len(parse_subjects(subjects[p])))
    oof["n_fields"] = oof.weights.map(len)
    oof["allowed"] = oof.project_id.map(
        lambda p: allowed_with_secondary(subjects[p], primary, secondary))

    if (oof.n_fields == 0).any():
        n = int((oof.n_fields == 0).sum())
        print(f"warning: {n} projects have no mappable subject and are excluded")
        oof = oof[oof.n_fields > 0].reset_index(drop=True)

    # --- how multi-label is the reference standard --------------------------
    print(f"=== the gold standard's own structure (n = {len(oof)}) ===")
    print("research subjects per project:")
    print(oof.n_subjects.value_counts().sort_index().to_string())
    print("\ndistinct fields per project after the crosswalk:")
    print(oof.n_fields.value_counts().sort_index().to_string())
    multi = oof.n_fields > 1
    print(f"\nspanning more than one field: {int(multi.sum())} of {len(oof)} "
          f"({multi.mean():.0%})")
    print(f"mean weight on the top field: "
          f"{oof.weights.map(lambda d: max(d.values())).mean():.3f}")

    # --- the four figures ----------------------------------------------------
    oof["strict"] = oof.pred_field == oof.true_field
    oof["any_tagged"] = [r.pred_field in r.weights for r in oof.itertuples()]
    oof["with_secondary"] = [r.pred_field in r.allowed for r in oof.itertuples()]
    oof["credit"] = [r.weights.get(r.pred_field, 0.0) for r in oof.itertuples()]
    ceiling = oof.weights.map(lambda d: max(d.values())).mean()

    print(f"\n=== accuracy under four definitions ===")
    print(f"  strict, top-1 == funder's top field        {oof.strict.mean():.1%}")
    print(f"  primary or crosswalk-declared secondary    {oof.with_secondary.mean():.1%}")
    print(f"  any field the funder tagged                {oof.any_tagged.mean():.1%}")
    print(f"\n  partial credit                             {oof.credit.mean():.3f}")
    print(f"    ceiling (always the funder's top field)  {ceiling:.3f}")
    print(f"    proportion of achievable credit          {oof.credit.mean()/ceiling:.1%}")

    print(f"\n  single-field projects (n={int((~multi).sum())}): "
          f"strict {oof[~multi].strict.mean():.1%}, credit {oof[~multi].credit.mean():.3f}")
    print(f"  multi-field projects  (n={int(multi.sum())}): "
          f"strict {oof[multi].strict.mean():.1%}, credit {oof[multi].credit.mean():.3f}, "
          f"any tagged {oof[multi].any_tagged.mean():.1%}")

    # --- rank profile, the most informative table ---------------------------
    def rank_of(row):
        order = sorted(row.weights, key=row.weights.get, reverse=True)
        return order.index(row.pred_field) + 1 if row.pred_field in order else 0

    oof["rank"] = [rank_of(r) for r in oof.itertuples()]
    print("\n=== where the model's choice sits in the funder's ranking ===")
    for k, v in oof["rank"].value_counts().sort_index().items():
        label = "not tagged at all" if k == 0 else f"the funder's #{k} field"
        print(f"  {label:24s} {v:>4}  ({v/len(oof):.0%})")

    summary = pd.DataFrame([
        dict(metric="strict", value=round(float(oof.strict.mean()), 4), n=len(oof)),
        dict(metric="primary_or_secondary",
             value=round(float(oof.with_secondary.mean()), 4), n=len(oof)),
        dict(metric="any_tagged_field",
             value=round(float(oof.any_tagged.mean()), 4), n=len(oof)),
        dict(metric="partial_credit", value=round(float(oof.credit.mean()), 4), n=len(oof)),
        dict(metric="partial_credit_ceiling", value=round(float(ceiling), 4), n=len(oof)),
        dict(metric="proportion_of_achievable",
             value=round(float(oof.credit.mean() / ceiling), 4), n=len(oof)),
        dict(metric="projects_spanning_multiple_fields",
             value=round(float(multi.mean()), 4), n=int(multi.sum())),
    ])
    RESULTS.mkdir(parents=True, exist_ok=True)
    summary.to_csv(RESULTS / "multilabel_evaluation.csv", index=False)
    oof.drop(columns=["weights", "allowed"]).to_csv(
        RESULTS / "multilabel_per_project.csv", index=False)
    print(f"\nWrote multilabel_evaluation.csv and multilabel_per_project.csv")
    print("Report strict first. The looser figures need the interdisciplinarity "
          "argument\nmade explicitly alongside them, not left implied.")


if __name__ == "__main__":
    main()
